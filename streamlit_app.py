import io
import re
import unicodedata
from decimal import Decimal, InvalidOperation

import pandas as pd
import streamlit as st
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import phonenumbers
from phonenumbers.phonenumberutil import country_code_for_region


# ------------------ UI / Page config ------------------
st.set_page_config(page_title="Lead Quality Checker", page_icon="✅", layout="wide")

HEADER_YELLOW = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
CELL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CELL_BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CELL_CHANGE = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # changed cell highlight


# ------------------ Normalisation helpers ------------------
def norm_text(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s.casefold()


def norm_picklist_value(s) -> str:
    """
    Strict normalisation for picklist comparisons only.
    Safe equivalences:
    - Case/spacing differences ignored
    - '&' treated same as 'and'
    - punctuation differences ignored
    """
    x = norm_text(s)
    x = x.replace("&", " and ")
    x = x.replace(",", " ")
    x = x.replace("/", " ")
    x = x.replace("-", " ")
    x = re.sub(r"\s+", " ", x).strip()
    return x


# ------------------ Flexible header detection ------------------
def _clean_header(h) -> str:
    h = "" if h is None else str(h)
    h = h.strip().lower()
    h = re.sub(r"\s+", " ", h)
    return h


def find_best_column(df: pd.DataFrame, keywords, banned=(), prefer_startswith=True):
    cols = list(df.columns)
    scored = []

    for c in cols:
        hc = _clean_header(c)

        if any(b in hc for b in banned):
            continue

        score = 0
        for kw in keywords:
            kw = kw.lower()
            if kw in hc:
                score += 10
            if prefer_startswith and hc.startswith(kw):
                score += 3

        score += max(0, 5 - len(hc.split()))
        if score > 0:
            scored.append((score, c))

    scored.sort(reverse=True, key=lambda x: x[0])
    return scored[0][1] if scored else None


def detect_columns(master_df: pd.DataFrame, pick_df: pd.DataFrame):
    return {
        # MASTER
        "master_company": find_best_column(master_df, ["companyname", "company name", "company"], banned=["domain", "website"]),
        "master_country": find_best_column(master_df, ["lead_country", "company country", "country"], banned=["county"]),
        "master_industry": find_best_column(master_df, ["industry", "c_industry"], banned=[]),
        "master_departments": find_best_column(master_df, ["department", "departments", "function"], banned=[]),
        "master_asset_title": find_best_column(master_df, ["asset_title", "asset title", "asset"], banned=[]),
        "master_phone": find_best_column(master_df, ["telephone", "phone", "tel", "mobile", "cell"], banned=["phonetic"]),
        "master_website": find_best_column(master_df, ["website", "domain", "web", "url"], banned=["email"]),
        "master_email": find_best_column(master_df, ["email"], banned=[]),
        "master_jobtitle": find_best_column(master_df, ["jobtitle", "job title", "title"], banned=[]),

        # NEW: Seniority picklist-based standardisation (Master → Picklist)
        "master_seniority": find_best_column(master_df, ["seniority", "senior"], banned=[]),

        # PICKLIST
        "pick_country": find_best_column(pick_df, ["lead_country", "country", "company country"], banned=["county"]),
        "pick_industry": find_best_column(pick_df, ["industry", "c_industry"], banned=[]),
        "pick_departments": find_best_column(pick_df, ["department", "departments", "function"], banned=[]),
        "pick_asset_title": find_best_column(pick_df, ["asset_title", "asset title", "asset"], banned=[]),

        # NEW: Seniority in picklist
        "pick_seniority": find_best_column(pick_df, ["seniority", "senior"], banned=[]),
    }


# ------------------ Country standardisation ------------------
COUNTRY_ALIASES_TO_CANON = {
    # UK -> United Kingdom
    "uk": "united kingdom",
    "u.k.": "united kingdom",
    "gb": "united kingdom",
    "great britain": "united kingdom",
    "britain": "united kingdom",
    "england": "united kingdom",
    "scotland": "united kingdom",
    "wales": "united kingdom",
    "northern ireland": "united kingdom",

    # US -> United States
    "us": "united states",
    "u.s.": "united states",
    "usa": "united states",
    "u.s.a.": "united states",
    "united states of america": "united states",
    "america": "united states",

    # UAE -> United Arab Emirates
    "uae": "united arab emirates",
    "u.a.e.": "united arab emirates",
    "dubai": "united arab emirates",
    "abu dhabi": "united arab emirates",

    # Common alternates
    "holland": "netherlands",
    "czech republic": "czechia",
    "russian federation": "russia",
    "viet nam": "vietnam",
    "republic of korea": "south korea",
    "korea, south": "south korea",
}


def canon_country_key(s) -> str:
    x = norm_text(s)
    return COUNTRY_ALIASES_TO_CANON.get(x, x)


COUNTRY_TO_REGIONS = {
    "united kingdom": ["GB"],
    "ireland": ["IE"],
    "united states": ["US"],
    "canada": ["CA"],
    "australia": ["AU"],
    "new zealand": ["NZ"],
    "singapore": ["SG"],
    "malaysia": ["MY"],
    "thailand": ["TH"],
    "indonesia": ["ID"],
    "philippines": ["PH"],
    "hong kong": ["HK"],
    "taiwan": ["TW"],
    "japan": ["JP"],
    "south korea": ["KR"],
    "china": ["CN"],
    "india": ["IN"],
    "vietnam": ["VN"],
    "france": ["FR"],
    "germany": ["DE"],
    "spain": ["ES"],
    "italy": ["IT"],
    "netherlands": ["NL"],
    "sweden": ["SE"],
    "norway": ["NO"],
    "denmark": ["DK"],
    "finland": ["FI"],
    "switzerland": ["CH"],
    "austria": ["AT"],
    "belgium": ["BE"],
    "portugal": ["PT"],
}


# ------------------ Company ↔ domain helpers (STRICT) ------------------
SUFFIXES = {
    "ltd","limited","co","company","corp","corporation","inc","incorporated",
    "plc","public","llc","lp","llp","ulc","pc","pllc","sa","ag","nv","se","bv",
    "oy","ab","aps","as","kft","zrt","rt","sarl","sas","spa","gmbh","ug",
    "kg","kgaa","pte","pty","sdn","bhd","kk","k.k.","co.","ltd.","inc.","plc.",
    "holdings","holding","group"
}


def _clean_domain(domain: str) -> str:
    if not isinstance(domain, str):
        return ""
    domain = domain.lower().strip()
    domain = re.sub(r"^https?://", "", domain)
    domain = re.sub(r"/.*$", "", domain)
    domain = domain.replace("www.", "")
    return domain.strip()


def _domain_base(domain: str) -> str:
    d = _clean_domain(domain)
    if not d:
        return ""
    base = d.split(".")[0]  # STRICT
    base = re.sub(r"[^a-z0-9]", "", base.lower())
    return base


COUNTRY_WORDS_FOR_ACRONYM = {
    "uk", "u.k", "gb", "great", "britain", "england", "scotland", "wales", "ireland",
    "usa", "us", "u.s", "united", "states", "america",
    "uae", "u.a.e", "dubai", "abu", "dhabi",
}
STOPWORDS = {"of", "and", "the", "for", "to", "a"}


def _company_tokens(company: str) -> list[str]:
    if not isinstance(company, str):
        return []
    s = company.strip()
    s = re.sub(r"^\s*uk\s*[-–—:]\s*", "", s, flags=re.IGNORECASE)  # strip leading "UK -"
    s = re.sub(r"[^A-Za-z0-9\s]", " ", s)
    toks = [t.lower() for t in s.split() if t.strip()]

    # remove suffixes, but keep "group"
    toks = [t for t in toks if (t not in SUFFIXES) or (t == "group")]
    toks = [t for t in toks if t not in STOPWORDS]
    toks = [t for t in toks if t not in COUNTRY_WORDS_FOR_ACRONYM]
    return toks


def _company_acronym(company: str) -> str:
    toks = _company_tokens(company)
    if not toks:
        return ""
    return "".join(t[0] for t in toks if t).upper()


def _is_subsequence(short: str, long: str) -> bool:
    it = iter(long)
    return all(ch in it for ch in short)


def compare_company_domain(company, domain) -> tuple[str, int, str]:
    c_raw = company if isinstance(company, str) else ""
    d_raw = domain if isinstance(domain, str) else ""

    if not c_raw or not d_raw:
        return "Unsure – Please Check", 0, "missing values"

    dbase = _domain_base(d_raw)
    if not dbase:
        return "Unsure – Please Check", 0, "missing/invalid domain"

    acr = _company_acronym(c_raw)
    if acr and dbase == acr.lower():
        return "Likely Match", 99, "company acronym equals domain"

    c_compact = re.sub(r"[^A-Za-z0-9]", "", c_raw).upper()
    if 2 <= len(c_compact) <= 6 and c_compact.isalpha():
        if _is_subsequence(c_compact.lower(), dbase):
            return "Likely Match", 92, "company abbreviation in domain"

    toks = _company_tokens(c_raw)
    joined = "".join(toks)
    if joined and joined in dbase:
        return "Likely Match", 95, "company tokens contained in domain"

    token_str = " ".join(toks)
    score = int(max(
        fuzz.token_sort_ratio(token_str, dbase),
        fuzz.partial_ratio(token_str, dbase)
    ))
    if score >= 85:
        return "Likely Match", score, "strong fuzzy"
    if score >= 70:
        return "Unsure – Please Check", score, "weak fuzzy"
    return "Likely NOT Match", score, "low similarity"


# ------------------ Seniority parsing (derived; NOT picklist overwrite) ------------------
def parse_seniority(title):
    # derived classification from job title text
    if not isinstance(title, str):
        return "Entry"

    t = norm_text(title)

    if re.search(r"\b(chief|ceo|cfo|cio|cto|coo|chro|president)\b", t):
        return "C Suite"
    if re.search(r"\b(senior vice president|svp|vice president|vp)\b", t):
        return "VP"
    if re.search(r"\bhead\b", t):
        return "Head"
    if re.search(r"\bdirector\b", t):
        return "Director"
    if re.search(r"\bmanager\b", t):
        return "Manager"

    return "Entry"


# ------------------ Phone cleaning + validation ------------------
def phone_to_string(raw_phone) -> str:
    if raw_phone is None:
        return ""
    s = str(raw_phone).strip()
    if s == "":
        return ""

    if "e" in s.lower():
        try:
            s = format(int(Decimal(s)), "d")
        except (InvalidOperation, ValueError):
            pass

    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]

    return s


def normalise_phone_for_check(raw_phone: str, region: str) -> str:
    phone = phone_to_string(raw_phone)
    phone = re.sub(r"[^\d+]", "", phone)

    if phone.startswith("00"):
        phone = "+" + phone[2:]
    if phone.startswith("+"):
        return phone

    try:
        cc = str(country_code_for_region(region))
    except Exception:
        cc = ""

    if cc and phone.startswith(cc):
        return "+" + phone

    return phone


def phone_country_check(raw_phone, country_label) -> tuple[str, str]:
    if raw_phone is None or str(raw_phone).strip() == "":
        return "Unsure", "Missing phone"
    if country_label is None or str(country_label).strip() == "":
        return "Unsure", "Missing country"

    canon = canon_country_key(country_label)
    regions = COUNTRY_TO_REGIONS.get(canon, [])
    if not regions:
        return "Unsure", f"No phone region mapping for '{country_label}'"

    for region in regions:
        phone_for_check = normalise_phone_for_check(raw_phone, region)
        try:
            num = phonenumbers.parse(phone_for_check, region)
            actual_region = phonenumbers.region_code_for_number(num) or ""

            # Treat NANP (US/CA) as acceptable either way
            expected_set = set(regions)
            nanp_set = {"US", "CA"}

            if actual_region and actual_region not in expected_set:
                if expected_set.issubset(nanp_set) and actual_region in nanp_set:
                    # allow CA when expecting US, and US when expecting CA
                    pass
                else:
                    return "Warning", f"Parsed region {actual_region} != expected {regions}"

            if not phonenumbers.is_possible_number(num):
                return "Warning", "Number not possible for country"
            if not phonenumbers.is_valid_number(num):
                return "Warning", "Number not valid for country"

            return "Match", "Valid for country"
        except Exception:
            continue
    return "Unsure", "Could not parse phone for supplied country"


# ------------------ Main processing ------------------
def run_matching(master_bytes: bytes, picklist_bytes: bytes, apply_colours: bool, progress_cb=None) -> bytes:
    df_master = pd.read_excel(io.BytesIO(master_bytes), dtype=str, keep_default_na=False)
    df_picklist = pd.read_excel(io.BytesIO(picklist_bytes), dtype=str, keep_default_na=False)

    df_out = df_master.copy()
    added_cols = []

    colmap = detect_columns(df_master, df_picklist)

    # Build picklist "source of truth" country labels
    picklist_country_label_by_canon = {}
    pick_country_col = colmap.get("pick_country")
    if pick_country_col:
        for v in df_picklist[pick_country_col]:
            v = str(v).strip()
            if v:
                picklist_country_label_by_canon[canon_country_key(v)] = v

    # These are the columns we will standardise:
    # - match master → picklist using normalised key
    # - overwrite master value ONLY when there is an exact normalised match
    EXACT_PAIRS = [
        ("country", colmap.get("master_country"), colmap.get("pick_country")),
        ("industry", colmap.get("master_industry"), colmap.get("pick_industry")),
        ("departments", colmap.get("master_departments"), colmap.get("pick_departments")),
        ("asset_title", colmap.get("master_asset_title"), colmap.get("pick_asset_title")),
        ("seniority", colmap.get("master_seniority"), colmap.get("pick_seniority")),
    ]

    # Store ORIGINAL values before overwrites (for change notes)
    original_series_by_friendly = {}
    for friendly, master_col, pick_col in EXACT_PAIRS:
        if master_col and master_col in df_master.columns:
            original_series_by_friendly[friendly] = df_master[master_col].astype(str).copy()

    if progress_cb:
        progress_cb(0.15, "Running picklist checks...")

    for friendly, master_col, pick_col in EXACT_PAIRS:
        out_col = f"Match_{friendly}"
        df_out[out_col] = "Column Missing"
        added_cols.append(out_col)

        if not master_col or not pick_col:
            continue

        # Picklist map: normalised key -> exact picklist text (bible formatting)
        pick_map = {
            norm_picklist_value(v): str(v).strip()
            for v in df_picklist[pick_col]
            if str(v).strip() != ""
        }

        matches = []
        new_vals = []

        for raw_val in df_master[master_col]:
            v = str(raw_val)

            # Special handling: country canon -> force to picklist format where possible
            if friendly == "country":
                canon = canon_country_key(v)
                if canon in picklist_country_label_by_canon:
                    v = picklist_country_label_by_canon[canon]

            key = norm_picklist_value(v)

            # ONLY overwrite when we are sure: exact match on normalised key
            if key in pick_map:
                matches.append("Yes")
                new_vals.append(pick_map[key])  # overwrite to picklist formatting
            else:
                matches.append("No")
                new_vals.append(v)  # keep original

        df_out[out_col] = matches
        df_out[master_col] = new_vals

    # Change notes for every standardised field (Country/Seniority/etc.)
    if progress_cb:
        progress_cb(0.30, "Auditing changes...")

    for friendly, master_col, pick_col in EXACT_PAIRS:
        if not master_col or friendly not in original_series_by_friendly:
            continue

        orig_vals = original_series_by_friendly[friendly]
        final_vals = df_out[master_col].astype(str)

        notes = []
        for o, f in zip(orig_vals, final_vals):
            o2 = str(o).strip()
            f2 = str(f).strip()
            if o2 and f2 and norm_text(o2) != norm_text(f2):
                notes.append(f"{o2} → {f2}")
            else:
                notes.append("")

        note_col = f"{friendly.title()}_Change_Note"
        df_out[note_col] = notes
        added_cols.append(note_col)

    # Derived seniority from Job Title (separate from picklist seniority)
    if progress_cb:
        progress_cb(0.45, "Parsing seniority...")

    jobtitle_col = colmap.get("master_jobtitle")
    df_out["Parsed_Seniority"] = df_out[jobtitle_col].apply(parse_seniority) if jobtitle_col else ""
    added_cols.append("Parsed_Seniority")

    # Company/domain check
    if progress_cb:
        progress_cb(0.60, "Checking company vs domain...")

    company_col = colmap.get("master_company")
    website_col = colmap.get("master_website")
    email_col = colmap.get("master_email")

    statuses, scores, reasons = [], [], []
    if company_col:
        for i in range(len(df_master)):
            comp = str(df_master.at[i, company_col])

            dom = ""
            if website_col and str(df_master.at[i, website_col]).strip():
                dom = str(df_master.at[i, website_col]).strip()
            elif email_col and str(df_master.at[i, email_col]).strip():
                em = str(df_master.at[i, email_col])
                if "@" in em:
                    dom = em.split("@", 1)[1].strip()

            status, score, reason = compare_company_domain(comp, dom)
            statuses.append(status)
            scores.append(score)
            reasons.append(reason)
    else:
        statuses = ["company column not found"] * len(df_master)
        scores = [0] * len(df_master)
        reasons = ["company column not found"] * len(df_master)

    df_out["Company_Domain_Status"] = statuses
    df_out["Company_Domain_Score"] = scores
    df_out["Company_Domain_Reason"] = reasons
    added_cols += ["Company_Domain_Status", "Company_Domain_Score", "Company_Domain_Reason"]

    # Phone vs country check
    if progress_cb:
        progress_cb(0.75, "Checking phone vs country...")

    phone_col = colmap.get("master_phone")
    master_country_col = colmap.get("master_country")

    if phone_col and master_country_col:
        st_col = f"{phone_col}_PhoneCountry_Status"
        rs_col = f"{phone_col}_PhoneCountry_Reason"
        added_cols += [st_col, rs_col]

        out_status, out_reason = [], []
        for i in range(len(df_master)):
            raw_phone = df_master.at[i, phone_col]
            ctry = df_out.at[i, master_country_col] if master_country_col in df_out.columns else ""
            s, r = phone_country_check(raw_phone, ctry)
            out_status.append(s)
            out_reason.append(r)

        df_out[st_col] = out_status
        df_out[rs_col] = out_reason
    else:
        df_out["PhoneCountry_Status"] = "phone or country column not found"
        df_out["PhoneCountry_Reason"] = ""
        added_cols += ["PhoneCountry_Status", "PhoneCountry_Reason"]

    # Required fields (basic)
    if progress_cb:
        progress_cb(0.85, "Checking required fields...")

    required_cols = []
    if email_col:
        required_cols.append(email_col)
    if company_col:
        required_cols.append(company_col)
    if master_country_col:
        required_cols.append(master_country_col)

    if required_cols:
        missing_list = []
        for i in range(len(df_out)):
            missing = []
            for c in required_cols:
                v = df_out.at[i, c]
                if v is None or str(v).strip() == "":
                    missing.append(c)
            missing_list.append(", ".join(missing))
        df_out["Missing_Required_Fields"] = missing_list
        added_cols.append("Missing_Required_Fields")

    # Overall status + issues summary
    if progress_cb:
        progress_cb(0.92, "Building overall status...")

    match_cols = [c for c in df_out.columns if c.startswith("Match_")]
    phone_status_cols = [c for c in df_out.columns if c.endswith("_PhoneCountry_Status")]
    if "PhoneCountry_Status" in df_out.columns:
        phone_status_cols.append("PhoneCountry_Status")

    good_company_status = {"Likely Match", "Match"}

    overall_status, overall_issues = [], []
    for i in range(len(df_out)):
        issues = []

        for c in match_cols:
            if norm_text(df_out.at[i, c]) != "yes":
                issues.append(c)

        if "Company_Domain_Status" in df_out.columns:
            if str(df_out.at[i, "Company_Domain_Status"]).strip() not in good_company_status:
                issues.append("Company_Domain_Status")

        for c in phone_status_cols:
            if str(df_out.at[i, c]).strip() != "Match":
                issues.append(c)

        if "Missing_Required_Fields" in df_out.columns:
            if str(df_out.at[i, "Missing_Required_Fields"]).strip():
                issues.append("Missing_Required_Fields")

        if issues:
            overall_status.append("REVIEW")
            overall_issues.append("; ".join(issues))
        else:
            overall_status.append("PASS")
            overall_issues.append("")

    df_out["Overall_Status"] = overall_status
    df_out["Overall_Issues"] = overall_issues
    added_cols += ["Overall_Status", "Overall_Issues"]

    # Write Excel
    if progress_cb:
        progress_cb(0.96, "Writing Excel output...")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Results")
    output.seek(0)

    if not apply_colours:
        return output.read()

    # ------------------ Apply formatting ------------------
    wb = load_workbook(output)
    ws = wb["Results"]

    headers = [cell.value for cell in ws[1]]
    col_index_raw = {("" if v is None else str(v)): i + 1 for i, v in enumerate(headers)}
    col_index_clean = {_clean_header(v): i + 1 for i, v in enumerate(headers)}
    max_row = ws.max_row

    def get_col_idx(name: str):
        if name is None:
            return None
        name_raw = str(name)
        if name_raw in col_index_raw:
            return col_index_raw[name_raw]
        return col_index_clean.get(_clean_header(name_raw))

    # Yellow headers for added columns
    for col_name in set(added_cols):
        cidx = get_col_idx(col_name)
        if cidx:
            ws.cell(row=1, column=cidx).fill = HEADER_YELLOW

    def fill_column(col_name: str, is_green_fn):
        cidx = get_col_idx(col_name)
        if not cidx:
            return
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=cidx).value
            if val is None:
                continue
            ws.cell(row=r, column=cidx).fill = CELL_GREEN if is_green_fn(str(val)) else CELL_BLUE

    # Match_* columns
    for mc in match_cols:
        fill_column(mc, lambda v: norm_text(v) == "yes")

    # Company domain status
    fill_column("Company_Domain_Status", lambda v: norm_text(v) in {"likely match", "match"})

    # Phone status columns
    for pc in phone_status_cols:
        fill_column(pc, lambda v: norm_text(v) == "match")

    # Phone reason columns
    phone_reason_cols = [c for c in df_out.columns if c.endswith("_PhoneCountry_Reason")]
    if "PhoneCountry_Reason" in df_out.columns:
        phone_reason_cols.append("PhoneCountry_Reason")
    for rc in phone_reason_cols:
        fill_column(rc, lambda v: norm_text(v) == "valid for country")

    # Highlight cells that were changed based on *_Change_Note columns
    change_note_cols = [c for c in df_out.columns if c.endswith("_Change_Note")]
    for note_col in change_note_cols:
        note_idx = get_col_idx(note_col)
        if not note_idx:
            continue

        friendly = note_col.replace("_Change_Note", "").lower()
        base_master_col = None
        for fr, master_col, pick_col in EXACT_PAIRS:
            if fr == friendly:
                base_master_col = master_col
                break

        base_idx = get_col_idx(base_master_col) if base_master_col else None
        if not base_idx:
            continue

        for r in range(2, max_row + 1):
            note_val = ws.cell(row=r, column=note_idx).value
            if note_val is not None and str(note_val).strip() != "":
                ws.cell(row=r, column=base_idx).fill = CELL_CHANGE

    # Missing required fields
    if get_col_idx("Missing_Required_Fields"):
        cidx = get_col_idx("Missing_Required_Fields")
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=cidx).value
            if val is None or str(val).strip() == "":
                ws.cell(row=r, column=cidx).fill = CELL_GREEN
            else:
                ws.cell(row=r, column=cidx).fill = CELL_BLUE

    # Overall status + issues
    fill_column("Overall_Status", lambda v: norm_text(v) == "pass")

    if get_col_idx("Overall_Issues"):
        cidx = get_col_idx("Overall_Issues")
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=cidx).value
            if val is None or str(val).strip() == "":
                ws.cell(row=r, column=cidx).fill = CELL_GREEN
            else:
                ws.cell(row=r, column=cidx).fill = CELL_BLUE

    out2 = io.BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.read()


# ------------------ Streamlit UI ------------------
st.markdown("## Lead Quality Checker")
st.caption(
    "Upload the **Lead Master** and the **Picklist**, then click **Run matching**.\n\n"
    "🟩 Green = OK / match, 🟦 Blue = review.\n"
    "📞 Phone validation runs in the background and does NOT change your phone field.\n"
    "🌍 Country names are standardised to the picklist format where possible (e.g. GB/England → United Kingdom).\n"
    "🧩 Picklist matching treats '&' and 'and' as equivalent (e.g. Travel & Tourism).\n"
    "📌 Any master value overwritten to match picklist formatting is highlighted."
)

col1, col2 = st.columns(2)
with col1:
    master_file = st.file_uploader("Upload Lead Master (.xlsx)", type=["xlsx"], key="master")
with col2:
    picklist_file = st.file_uploader("Upload Picklist (.xlsx)", type=["xlsx"], key="picklist")

apply_colours = st.toggle("Colour-code results (green/blue) + yellow headers for new columns", value=True)

if "last_output_bytes" not in st.session_state:
    st.session_state.last_output_bytes = None

run_btn = st.button(
    "▶ Run matching",
    type="primary",
    use_container_width=True,
    disabled=not (master_file and picklist_file),
)

if run_btn:
    progress = st.progress(0.0, text="Starting...")

    def prog(p, text=""):
        progress.progress(min(max(float(p), 0.0), 1.0), text=text)

    try:
        output_bytes = run_matching(
            master_file.read(),
            picklist_file.read(),
            apply_colours=apply_colours,
            progress_cb=prog
        )
        st.session_state.last_output_bytes = output_bytes
        st.success("Processing complete.")
    except Exception as e:
        st.session_state.last_output_bytes = None
        st.error(f"Error: {e}")

if st.session_state.last_output_bytes:
    st.download_button(
        label="⬇ Download Processed File",
        data=st.session_state.last_output_bytes,
        file_name="Full_Check_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
